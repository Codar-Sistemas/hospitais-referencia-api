"""
Dual-format workbook reader for the CGCAN files.

The main oncology file is a legacy .xls (BIFF / OLE2 Compound Document)
shipped with a misleading .xlsx extension — openpyxl cannot open it, xlrd
can ONLY open it. Format is sniffed from magic bytes, never from the file
name, so a future re-export as true XLSX is absorbed transparently.

Both paths normalize to the same shape: sheet name -> grid of cell values
(`None` for empty cells), which the layout mappers consume.
"""

from __future__ import annotations

from io import BytesIO
from typing import Literal

import xlrd
from openpyxl import load_workbook

from scripts.shared.spreadsheets import SheetLayoutError

# OLE2 Compound Document (legacy .xls) and ZIP (xlsx) signatures.
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_ZIP_MAGIC = b"PK\x03\x04"


def detect_format(content: bytes) -> Literal["xls", "xlsx"]:
    if content.startswith(_OLE2_MAGIC):
        return "xls"
    if content.startswith(_ZIP_MAGIC):
        return "xlsx"
    raise SheetLayoutError(f"Unrecognized workbook format (magic bytes {content[:8].hex()!r})")


def read_sheets(content: bytes, source_url: str) -> dict[str, list[list[object]]]:
    """Returns {sheet_name: grid}; empty cells come through as None in both
    formats (xlrd renders empty merged cells as '' — normalized here)."""
    fmt = detect_format(content)
    if fmt == "xls":
        return _read_xls(content)
    return _read_xlsx(content, source_url)


def _read_xls(content: bytes) -> dict[str, list[list[object]]]:
    book = xlrd.open_workbook(file_contents=content)
    sheets: dict[str, list[list[object]]] = {}
    for sheet in book.sheets():
        grid: list[list[object]] = []
        for r in range(sheet.nrows):
            row: list[object] = []
            for c in range(sheet.ncols):
                value = sheet.cell_value(r, c)
                row.append(None if value == "" else value)
            grid.append(row)
        sheets[sheet.name] = grid
    return sheets


def _read_xlsx(content: bytes, source_url: str) -> dict[str, list[list[object]]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        return {
            sheet.title: [list(row) for row in sheet.iter_rows(values_only=True)]
            for sheet in workbook.worksheets
        }
    except Exception as e:  # corrupt zip member etc.
        raise SheetLayoutError(f"Failed to read XLSX from {source_url}: {e}") from e
    finally:
        workbook.close()
