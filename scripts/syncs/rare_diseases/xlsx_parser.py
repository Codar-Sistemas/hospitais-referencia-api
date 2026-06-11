"""
Parser for the gov.br rare-diseases XLSX files.

Both national files share one layout (8 columns):
  UF | Município | Região | CNES | Estabelecimento | Tipo da Habilitação
     | Ano da Habilitação | Códigos Habilitados

Quirks handled here:
  - Columns are mapped BY HEADER NAME, not position, so a reordered
    column in a future upload does not silently corrupt the data.
  - An establishment qualified under several codes spans multiple rows:
    the first row carries all fields, follow-up rows carry only the
    "Códigos Habilitados" cell. Those continuation rows are folded into
    the parent row's `qualification_codes`.
  - Fully empty rows are skipped; CNES cells may arrive as int/float
    (Excel numeric cells) and are normalized to a zero-padded string.
"""

from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from scripts.syncs.rare_diseases.types import RareDiseaseXlsxRow

# Header cell -> RareDiseaseXlsxRow field. Keys are normalized (lowercase,
# accent-stripped) so cosmetic header edits don't break the mapping.
_EXPECTED_HEADERS = {
    "uf": "state_code",
    "municipio": "city",
    "regiao": "region",
    "cnes": "cnes",
    "estabelecimento": "name",
    "tipo da habilitacao": "qualification_type_raw",
    "ano da habilitacao": "qualification_year",
    "codigos habilitados": "qualification_codes",
}

_UF_RE = re.compile(r"^[A-Z]{2}$")


class XlsxLayoutError(RuntimeError):
    """The spreadsheet no longer matches the layout this parser expects.
    Raised loudly so a silent gov.br format change never half-applies."""


def parse_xlsx(content: bytes, source_url: str) -> list[RareDiseaseXlsxRow]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)

        columns = _locate_header(rows_iter, source_url)

        parsed: list[RareDiseaseXlsxRow] = []
        for raw in rows_iter:
            values = {field: raw[idx] if idx < len(raw) else None for field, idx in columns.items()}

            uf = _clean(values.get("state_code"))
            cnes_raw = values.get("cnes")
            code = _clean(values.get("qualification_codes"))

            if uf is None and cnes_raw is None:
                if code is None:
                    continue  # fully empty spacer row
                # Continuation row: only the codes column is filled.
                if not parsed:
                    raise XlsxLayoutError(
                        f"Continuation row with no parent establishment in {source_url}"
                    )
                parsed[-1]["qualification_codes"].append(code)
                continue

            if uf is None or not _UF_RE.match(uf.upper()):
                raise XlsxLayoutError(f"Unexpected UF value {uf!r} in {source_url}")

            name = _clean(values.get("name"))
            city = _clean(values.get("city"))
            qualification_type = _clean(values.get("qualification_type_raw"))
            if not name or not city or not qualification_type:
                raise XlsxLayoutError(
                    f"Establishment row missing name/city/type in {source_url}: {values!r}"
                )

            parsed.append(
                {
                    "state_code": uf.upper(),
                    "city": city,
                    "region": _clean(values.get("region")),
                    "cnes": _normalize_cnes(cnes_raw, source_url),
                    "name": name,
                    "qualification_type_raw": qualification_type,
                    "qualification_year": _as_year(values.get("qualification_year")),
                    "qualification_codes": [code] if code else [],
                    "source_url": source_url,
                }
            )
        return parsed
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
def _locate_header(rows_iter: Any, source_url: str) -> dict[str, int]:
    """Advance the row iterator past the header and return field->column-index.
    The header is the first row containing both a 'uf' and a 'cnes' cell —
    position-independent so a reordered upload still maps correctly."""
    for raw in rows_iter:
        normalized = [_normalize_header(cell) for cell in raw]
        if "uf" in normalized and "cnes" in normalized:
            columns: dict[str, int] = {}
            for idx, header in enumerate(normalized):
                field = _EXPECTED_HEADERS.get(header or "")
                if field:
                    columns[field] = idx
            missing = set(_EXPECTED_HEADERS.values()) - set(columns)
            # `region` is informational — tolerate its absence.
            missing.discard("region")
            if missing:
                raise XlsxLayoutError(
                    f"Header row in {source_url} is missing columns: {sorted(missing)}"
                )
            return columns
    raise XlsxLayoutError(f"No header row found in {source_url}")


def _normalize_header(value: Any) -> str | None:
    text = _clean(value)
    if text is None:
        return None
    stripped = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", stripped).strip().lower()


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_cnes(value: Any, source_url: str) -> str:
    """CNES is 7 digits by spec; Excel may deliver it as int, float or
    string with/without leading zeros."""
    if value is None:
        raise XlsxLayoutError(f"Establishment row without CNES in {source_url}")
    if isinstance(value, float):
        value = int(value)
    digits = re.sub(r"\D", "", str(value))
    if not digits or len(digits) > 7:
        raise XlsxLayoutError(f"Invalid CNES {value!r} in {source_url}")
    return digits.zfill(7)


def _as_year(value: Any) -> int | None:
    if value is None:
        return None
    try:
        year = int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None
    return year if 1990 <= year <= 2100 else None
